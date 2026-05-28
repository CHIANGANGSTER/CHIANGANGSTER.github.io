#!/usr/bin/env python3
"""Import E-column interior prompts from the local XLSX into prompts.json."""

from __future__ import annotations

import json
import shutil
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "_imports" / "prompt-input.xlsx"
PROMPTS = ROOT / "assets" / "data" / "prompts.json"
ARCHIVE = ROOT / "_imports" / "prompts-before-xlsx-import-2026-05-18.json"


GENERIC_TITLES = {
    "家居设计",
    "家居电商场景",
    "小红书风格家居",
    "室内场景",
}

STYLE_KEYWORDS = [
    ("新中式", ["新中式", "中式", "东亚美学", "书法", "茶具", "纸灯笼"]),
    ("日式", ["日式", "侘寂", "榻榻米", "纸灯笼", "茶壶"]),
    ("美式", ["美式", "格子垫", "木梁", "壁炉", "花图案"]),
    ("法式", ["法式", "石膏线", "法式壁炉"]),
    ("欧式", ["欧式", "大理石壁炉", "金色壁灯", "半身雕像"]),
    ("工业风", ["工业", "混凝土", "外露横梁", "黑色橱柜", "黑铁", "水泥"]),
    ("中古风", ["中古", "复古", "美拉德", "皮革", "胡桃", "棕色"]),
    ("奶油风", ["奶油", "米色", "奶白", "柔和", "治愈"]),
    ("原木风", ["原木", "木质", "木制", "浅色木", "木地板", "自然"]),
    ("轻奢风", ["轻奢", "豪华吊灯", "金色壁灯", "银色烛台", "烛台"]),
    ("北欧风", ["北欧", "浅木", "通风", "简洁"]),
    ("现代简约", ["现代简约", "现代", "简约", "极简", "时尚", "黑白灰"]),
    ("写实场景", ["写实", "真实", "CGI", "小红书风格"]),
]

SPACE_KEYWORDS = [
    ("客餐厅", ["客餐厅", "开放空间", "厨房和用餐区", "客厅和用餐区"]),
    ("客厅", ["客厅", "起居", "沙发", "电视", "娱乐柜"]),
    ("餐厅", ["餐厅", "餐桌", "用餐", "咖啡馆"]),
    ("厨房", ["厨房", "橱柜", "炉子", "岛台", "烤箱", "水槽"]),
    ("书房", ["书房", "书架", "电脑", "办公", "工作"]),
    ("卧室", ["卧室", "床", "床头"]),
    ("玄关", ["玄关", "入口", "入户"]),
    ("休闲角", ["休闲", "阅读角", "躺椅", "扶手椅"]),
    ("墙柜", ["墙柜", "搁板", "收纳柜", "展示柜"]),
]

MATERIAL_KEYWORDS = [
    ("木质", ["木质", "木制", "原木", "木地板", "胡桃"]),
    ("皮革", ["皮革", "皮沙发", "皮椅"]),
    ("大理石", ["大理石", "石材"]),
    ("金属", ["金属", "镀铬", "不锈钢", "黑铁"]),
    ("布艺", ["布艺", "亚麻", "织物", "毛绒"]),
    ("绿植", ["绿植", "植物", "盆栽", "盆景"]),
]

FEATURE_KEYWORDS = [
    ("暖色墙柜", ["温暖的木质装饰墙柜", "墙柜", "搁板"]),
    ("米色沙发", ["米色沙发", "白色布艺沙发", "大型白色布艺沙发"]),
    ("黑色皮革椅", ["黑色皮革扶手椅", "黑色皮革", "皮革扶手椅"]),
    ("抽象画客厅", ["大型抽象画", "抽象画"]),
    ("开放客餐厅", ["用餐区", "开放", "客厅和用餐区"]),
    ("绿植客厅", ["大型室内植物", "盆栽", "绿色植物"]),
    ("吊灯客厅", ["吊灯", "几何形状", "球形"]),
    ("电视墙客厅", ["电视", "娱乐柜", "电视墙"]),
    ("壁炉客厅", ["壁炉", "壁炉架"]),
    ("木质餐厅", ["木桌", "木椅", "餐桌"]),
    ("茶室书房", ["茶具", "茶壶", "书房"]),
    ("厨房岛台", ["厨房", "木岛", "岛台"]),
    ("混凝土空间", ["混凝土", "外露横梁"]),
    ("展示书架", ["书架", "展示台", "陈列墙"]),
    ("自然光客厅", ["自然光", "大窗户", "窗帘"]),
]

STYLE_ID = {
    "现代简约": "modern-minimal",
    "奶油风": "modern-cream",
    "中古风": "mid-century-modern",
    "原木风": "natural-wood",
    "轻奢风": "light-luxury",
    "工业风": "light-industrial",
    "法式": "french",
    "欧式": "european-classic",
    "美式": "american-classic",
    "日式": "japanese",
    "新中式": "new-chinese",
    "北欧风": "scandinavian",
    "写实场景": "realistic-interior",
}

STYLE_NAME = {
    "现代简约": "现代简约风格",
    "奶油风": "奶油风格",
    "中古风": "中古风格",
    "原木风": "原木风格",
    "轻奢风": "现代轻奢风格",
    "工业风": "轻工业风格",
    "法式": "法式风格",
    "欧式": "欧式古典风格",
    "美式": "美式风格",
    "日式": "日式风格",
    "新中式": "新中式风格",
    "北欧风": "北欧风格",
    "写实场景": "写实室内场景",
}

NEGATIVE_PROMPT = (
    "people, human figure, text, watermark, logo, duplicated furniture, "
    "distorted furniture legs, broken perspective, low resolution, blurry, "
    "overexposed highlights, messy clutter, dirty wall, inaccurate shadows"
)

IMAGE_PROMPT_SUFFIX = (
    "写实室内场景摄影，空间比例准确，家具结构清晰，材质真实，柔和自然光，"
    "高端家居目录图，干净构图，4:3。"
)


def pick_label(text: str, mapping: list[tuple[str, list[str]]], fallback: str) -> str:
    for label, keywords in mapping:
        if any(keyword in text for keyword in keywords):
            return label
    return fallback


def collect_materials(text: str) -> list[str]:
    return [
        label
        for label, keywords in MATERIAL_KEYWORDS
        if any(keyword in text for keyword in keywords)
    ][:2]


def is_generic_title(raw_title: str) -> bool:
    stripped = raw_title.strip()
    return not stripped or stripped in GENERIC_TITLES


def build_title(raw_title: str, prompt: str, used: Counter[str]) -> str:
    raw_title = raw_title.strip()
    if not is_generic_title(raw_title):
        base = raw_title[:24]
    else:
        text = f"{raw_title} {prompt}"
        style = pick_label(text, STYLE_KEYWORDS, "写实场景")
        space = pick_label(text, SPACE_KEYWORDS, "室内场景")
        feature = pick_label(text, FEATURE_KEYWORDS, space)
        if feature.endswith(space) or feature == space:
            base = f"{style}{feature}"
        else:
            base = f"{style}{feature}"

    used[base] += 1
    if used[base] == 1:
        return base
    return f"{base} {used[base]:02d}"


def build_image_prompt(prompt: str) -> str:
    return f"{prompt.strip()} {IMAGE_PROMPT_SUFFIX}"


def main() -> int:
    if not XLSX.exists():
        raise FileNotFoundError(XLSX)

    if not ARCHIVE.exists():
        shutil.copy2(PROMPTS, ARCHIVE)

    data = json.loads(PROMPTS.read_text(encoding="utf-8"))
    existing = [
        entry for entry in data.get("entries", [])
        if not str(entry.get("id", "")).startswith("xlsx-interior-")
    ]

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]
    imported = []
    used_titles: Counter[str] = Counter()

    for row in range(1, ws.max_row + 1):
        raw_title = str(ws.cell(row, 1).value or "").strip()
        source_name = str(ws.cell(row, 2).value or "").strip()
        industry = str(ws.cell(row, 3).value or "家居").strip()
        entry_type = str(ws.cell(row, 4).value or "室内场景").strip()
        prompt = str(ws.cell(row, 5).value or "").strip()
        if not prompt:
            continue

        text = f"{raw_title} {prompt}"
        style = pick_label(text, STYLE_KEYWORDS, "写实场景")
        space = pick_label(text, SPACE_KEYWORDS, "室内场景")
        tags = []
        for value in [entry_type, style, space, *collect_materials(text), "Excel导入"]:
            if value and value not in tags:
                tags.append(value)

        entry_id = f"xlsx-interior-{len(imported) + 1:03d}"
        imported.append({
            "id": entry_id,
            "title": build_title(raw_title, prompt, used_titles),
            "industry": industry or "家居",
            "type": entry_type or "室内场景",
            "styleId": STYLE_ID.get(style, "interior-reference"),
            "styleName": STYLE_NAME.get(style, style),
            "tags": tags,
            "image": "",
            "sourceImageName": source_name,
            "cmf": "",
            "prompt": prompt,
            "imagePrompt": build_image_prompt(prompt),
            "negativePrompt": NEGATIVE_PROMPT,
            "imageSource": "",
            "imageAspect": "4:3",
            "imageSize": "1344x1008",
        })

    data["version"] = "3-xlsx-interior-import"
    data["source"] = "PDF style-system + XLSX E-column interior prompts / generated by Codex"
    data["generatedAt"] = "2026-05-18"
    data.pop("archivedPreviousData", None)
    data["entries"] = existing + imported
    PROMPTS.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    duplicate_count = sum(count - 1 for count in used_titles.values() if count > 1)
    print(f"kept existing: {len(existing)}")
    print(f"imported: {len(imported)}")
    print(f"title suffixes added: {duplicate_count}")
    print(f"total: {len(data['entries'])}")
    print(f"first: {json.dumps(imported[0], ensure_ascii=False)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
